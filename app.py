import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from io import BytesIO
import pandas as pd
import datetime
import json
import re
from streamlit_gsheets import GSheetsConnection
import google.generativeai as genai

# SecretsからAPIキーを読み込む（設定されていない場合のエラー回避付き）
if "GEMINI_API_KEY" in st.secrets:
    genai.configure(api_key=st.secrets["GEMINI_API_KEY"])
    has_api_key = True
else:
    has_api_key = False
# --- 0. ページ設定 ---
st.set_page_config(page_title="保育指導計画システム", layout="wide", page_icon="📛")

# --- 1. 定数・データ定義 ---
TERMS = ["1期(4-5月)", "2期(6-8月)", "3期(9-12月)", "4期(1-3月)"]

# 定型文データ
TEIKEI_DATA = {
    "0歳児": {
        "健康": [
            "一人ひとりの生活リズムに合わせて心地よく過ごし、生理的欲求を満たす。",
            "離乳食を喜んで食べ、自分で手づかみ食べをしようとする。",
            "腹ばいやハイハイ、つかまり立ちをして、十分に体を動かそうとする。",
            "保育者にゆったりと抱かれ、安心して入眠する。",
            "沐浴や清拭を通して、体の清潔に保たれる心地よさを感じる。",
            "身の回りの物に興味を持ち、手を伸ばして掴もうとする。",
            "保育者と触れ合い遊びを楽しみ、声を出して笑う。",
            "戸外の空気に触れ、外の刺激を心地よく感じる。",
            "睡眠や食事の時間を一定に保ち、健康的な生活習慣を身につける。",
            "自分の手足を見つめたり動かしたりして、体の存在を認識する。"
        ],
        "人間関係": [
            "特定の保育者との関わりの中で、安心感と信頼感を持つ。",
            "あやされると笑ったり、声を出し返したりして応答を楽しむ。",
            "保育者の顔をじっと見つめ、表情を模倣しようとする。",
            "身近な大人に親しみを持ち、後追いや抱っこを求める。",
            "友達の存在に気づき、じっと見つめたり触れようとしたりする。",
            "自分の思いを泣き声やしぐさで保育者に伝えようとする。",
            "他児の泣き声に反応し、顔を覗き込もうとする。",
            "保育者の仲立ちによって、友達と同じ空間で過ごすことを楽しむ。",
            "名前を呼ばれると、振り向いたり笑顔を見せたりして応える。",
            "人見知りを経験しながら、特定の大人との絆を深めていく。"
        ],
        "環境": [
            "身近にある玩具に興味を持ち、舐める、叩く、振るなどして確かめる。",
            "音の鳴る玩具に反応し、自ら音を出して楽しもうとする。",
            "散歩中に見える草木や空の色など、自然の変化をじっと見つめる。",
            "動くものに興味を示し、目で追ったり手を伸ばしたりする。",
            "水の感触や土の匂いなど、五感を通して周囲の環境を感じる。",
            "身近な大人の持ち物に興味を持ち、触れようとする。",
            "鏡に映る自分の姿を見つめ、不思議そうに触れようとする。",
            "いないいないばあ等の遊びを通して、物の永続性に気づき始める。",
            "室内にある仕掛け玩具に触れ、繰り返し遊ぼうとする。",
            "戸外で鳥の声や風の音など、周囲の音に耳を傾ける。"
        ],
        "言葉": [
            "「アー」「ウー」などの喃語を発し、保育者とのやり取りを楽しむ。",
            "保育者の優しい語り掛けに耳を傾け、心地よさを感じる。",
            "絵本の絵を指差したり、保育者の読む声に反応したりする。",
            "自分の要求を声のトーンや強弱で使い分け、伝えようとする。",
            "音楽のリズムに合わせて、体を揺らしたり声を出したりする。",
            "「バイバイ」などの簡単な言葉と動作を、真似しようとする。",
            "身近な物の名前を聞いて、そちらの方を見ようとする。",
            "保育者の表情や声の調子から、相手の気持ちを感じ取ろうとする。",
            "一語文（「マンマ」「ブーブー」等）を話し、思いを伝えようとする。",
            "手遊び歌に合わせて、自分なりに手を動かそうとする。"
        ],
        "表現": [
            "保育者の歌声に合わせて、手足をバタバタさせて喜ぶ。",
            "シーツブランコや抱っこでの揺れを、全身で味わい表現する。",
            "いろいろな感触の布や紙に触れ、握ったり破いたりして遊ぶ。",
            "色のついた物や光るものに興味を持ち、じっと見つめる。",
            "砂を握ったり放したりして、その感触を自分なりに楽しむ。",
            "クレヨンなどを握り、紙に偶然色がつくことを喜ぶ。",
            "玩具を打ち鳴らし、リズムの面白さを感じようとする。",
            "食事中に食べ物を手で捏ねたり広げたりして、感触を確かめる。",
            "保育者のしぐさを真似て、パチパチやバイバイをする。",
            "周囲のいろいろな音に対し、自分なりの反応を見せる。"
        ]
    },
    "1歳児": {
        # 前回の回答の1歳児分を入れてください
    },
    "2歳児": {
        "健康": [
            "走る、跳ぶ、登るなどの運動を楽しみ、活発に体を動かす。",
            "保育者に見守られながら、自分で衣服を脱ごうとする。",
            "食事の前後には、保育者と一緒に手洗いをしようとする。",
            "スプーンやフォークを使って、自分で食べようとする意欲を持つ。",
            "尿意を意識し始め、保育者に伝えたりトイレに行こうとする。",
            "簡単な衣服の着脱（ズボンを上げる等）を自分で行おうとする。",
            "戸外で探索活動を楽しみ、体力を養う。",
            "鼻水が出ると保育者に知らせたり、自分で拭こうとしたりする。",
            "遊びと休息の切り替えをスムーズに行い、規則正しく過ごす。",
            "身の回りの危険なものに気づき、保育者の言葉に従って避ける。"
        ],
        "人間関係": [
            "保育者との安定した関係の中で、自分の思いを強く主張する。",
            "友達の持っている玩具を欲しがり、関わりを持とうとする。",
            "「貸して」「いいよ」などの言葉を使い、友達と遊ぼうとする。",
            "簡単なルールのある遊びを通して、友達と同じ目的を楽しむ。",
            "自分の好きな友達ができ、名前を呼んで一緒に遊ぼうとする。",
            "大人の真似をして、友達とおままごとやごっこ遊びを楽しむ。",
            "友達が困っている時に、心配そうに見つめたり近寄ったりする。",
            "集団での活動（手遊びやダンス等）を、友達と一緒に楽しむ。",
            "自分の持ち物を認識し、大切にしようとする気持ちが芽生える。",
            "保育者の助けを借りながら、順番を待とうとする。"
        ],
        "環境": [
            "動植物への興味が深まり、じっくり観察したり触れたりする。",
            "身近な自然物（どんぐりや石）を集め、自分なりに並べて遊ぶ。",
            "砂場や水遊びで、道具を使って形を作ったり運んだりする。",
            "身の回りの物の色や形の違いに気づき、分類しようとする。",
            "簡単な道具（糊やシール）を使い、自分なりに形にしようとする。",
            "散歩で見かける信号機や標識に興味を持ち、意味を知ろうとする。",
            "生活の中にある数（1、2、3等）に興味を持ち、数えようとする。",
            "身近な自然現象（雨、風、雷）に気づき、驚きや発見を共有する。",
            "自分のロッカーや靴箱の場所を覚え、進んで片付けようとする。",
            "積み木を高く積み上げたり、横に並べたりして構成を楽しむ。"
        ],
        "言葉": [
            "二語文や三語文を使い、自分の体験を保育者に話そうとする。",
            "「これ何？」と名前を尋ね、言葉の語彙を増やそうとする。",
            "簡単な絵本のストーリーを理解し、次の展開を期待して聞く。",
            "保育者や友達の問いかけに、自分の言葉で応答しようとする。",
            "自分の名前だけでなく、友達や保育者の名前も言おうとする。",
            "劇遊びの真似をして、役になりきった言葉を発しようとする。",
            "生活習慣に関する言葉（「いただきます」等）を自ら言う。",
            "保育者の歌う歌に合わせて、歌詞を口ずさむことを楽しむ。",
            "相手の言葉を聞き、自分の思いとの違いに気づき始める。",
            "好きな絵本を繰り返し読み、言葉の響きやリズムを楽しむ。"
        ],
        "表現": [
            "音楽に合わせて、動物の模倣をしたり自由な動きを楽しんだりする。",
            "クレヨンで丸や線を描き、それを何かに見立てて話そうとする。",
            "粘土を丸める、伸ばす、ちぎるなどの変化を楽しみ制作する。",
            "自分の経験したことを、絵や造形で表現しようとする。",
            "いろいろな色の絵の具を使い、色が混ざる面白さを味わう。",
            "空き箱を繋げたり色を塗ったりして、好きなものを作ろうとする。",
            "手遊びやダンスを覚え、友達と一緒に踊ることを喜ぶ。",
            "身近な大人やキャラクターになりきり、ごっこ遊びを広げる。",
            "スタンプ遊びを楽しみ、紙に模様ができる不思議さを感じる。",
            "出来上がった作品を、保育者や友達に嬉しそうに見せようとする。"
        ]
    },
    "3歳児": {
        "健康": [
            "運動遊びを通して、自分の体を思い切り動かすことを楽しむ。",
            "排泄を自立させ、自分から進んでトイレに行こうとする。",
            "衣服の着脱をほぼ一人で行い、脱いだものを畳もうとする。",
            "箸の使い方に興味を持ち、正しく持とうと意識する。",
            "食事の際、好き嫌いせずに何でも食べようとする意欲を持つ。",
            "手洗いやうがいの大切さを理解し、習慣化しようとする。",
            "健康への関心を持ち、自分の体の調子を保育者に伝える。",
            "戸外で活発に遊び、体力や持久力がついてくる。",
            "身の回りを清潔に保つ心地よさを感じ、進んで整理整頓する。",
            "午睡などで体を休める大切さを知り、静かに休息しようとする。"
        ],
        "人間関係": [
            "友達と共通の目的を持って、協力して遊ぼうとする。",
            "自分の思いを言葉で伝え、友達と折り合いをつけようとする。",
            "集団生活のルールを守り、順番や交代を意識して遊ぶ。",
            "困っている友達を助けたり、励ましたりする優しさが芽生える。",
            "保育者との関わりを楽しみつつ、友達同士の遊びを優先する。",
            "自分の気持ちをコントロールし、我慢したり譲ったりしようとする。",
            "友達と刺激し合いながら、新しい遊びに挑戦しようとする。",
            "クラスの一員であることを意識し、当番活動を頑張ろうとする。",
            "友達の良さに気づき、褒めたり認めたりしようとする。",
            "異年齢児との関わりを楽しみ、優しく接しようとする。"
        ],
        "環境": [
            "自然の不思議さに関心を持ち、図鑑などで調べようとする。",
            "栽培活動を通して、植物の生長を期待し世話を楽しもうとする。",
            "身の回りの物の性質（重い、軽い、浮く等）に興味を持つ。",
            "数や図形、文字に関心を持ち、生活の中で探そうとする。",
            "カレンダーや時計に興味を持ち、時間の流れを感じようとする。",
            "地域の施設（公園、図書館等）に親しみを持って利用する。",
            "廃材などを工夫して組み合わせ、自分のイメージを形にする。",
            "季節の行事の意味を知り、伝統的な遊びを体験しようとする。",
            "ゴミの分別に関心を持ち、身の回りを綺麗に保とうとする。",
            "散歩先で見つけた生き物の飼育に興味を持ち、観察を楽しむ。"
        ],
        "言葉": [
            "自分の経験したことや考えを、順序立てて話そうとする。",
            "相手の話を最後まで聞き、理解しようとする態度を持つ。",
            "新しい言葉や表現を使い、豊かな会話を楽しもうとする。",
            "文字に興味を持ち、自分の名前を読んだり書こうとしたりする。",
            "絵本のストーリーを記憶し、友達に読み聞かせようとする。",
            "「なぜ？」「どうして？」と質問を繰り返し、知識を広げる。",
            "友達とのトラブルを、言葉を使って解決しようと努める。",
            "劇遊びなどで、役に応じた言葉遣いを工夫して話す。",
            "しりとりや言葉遊びを楽しみ、言葉の響きに関心を深める。",
            "保育者の読み聞かせを静かに聞き、イメージを膨らませる。"
        ],
        "表現": [
            "音楽を聴いて、感じたことを体全体でダイナミックに表現する。",
            "自分の描きたいものを決め、形や色を工夫して描こうとする。",
            "ハサミや糊などの道具を正しく使い、複雑な制作に挑戦する。",
            "友達とイメージを共有し、役割を決めてごっこ遊びを展開する。",
            "いろいろな楽器に触れ、音色を楽しみながら合奏に親しむ。",
            "身近な素材を工夫し、役に必要な小道具を自作しようとする。",
            "発表会など、人前で表現することに自信と喜びを感じる。",
            "粘土や木切れなどを使い、立体的な作品を作ろうとする。",
            "色の濃淡や混色を楽しみ、自分の意図した色を作ろうとする。",
            "友達の表現した作品の良さに気づき、認め合おうとする。"
        ]
    },
    "4歳児": {
        "健康": [
            "ルールのある集団遊びを通して、力いっぱい体を動かす。",
            "自分の体の健康に関心を持ち、健康的な生活習慣を意識する。",
            "自分の体格に合った運動用具を使い、技術を身につけようとする。",
            "食事の栄養バランスに関心を持ち、進んで何でも食べる。",
            "身の回りの危険を予測し、安全な遊び方を自ら考える。",
            "避難訓練の重要性を理解し、迅速かつ冷静に行動しようとする。",
            "衣服の整理や始末を丁寧に行い、生活環境を整える。",
            "手洗い、うがい、換気などの感染予防を自ら進んで行う。",
            "休息と活動のバランスを自分で調整しようと意識する。",
            "体の仕組み（骨、筋肉等）に興味を持ち、大切にしようとする。"
        ],
        "人間関係": [
            "友達と意見を出し合い、共通の目標に向かって協力する。",
            "集団の中での自分の役割を理解し、責任を持って当番活動を行う。",
            "友達とのトラブルを、自分たちで話し合って解決しようとする。",
            "公共の場でのルールやマナーを守り、規律ある行動をとる。",
            "友達の失敗を許したり、励まし合ったりする仲間意識を持つ。",
            "異年齢の子供に対して思いやりを持ち、お世話を楽しもうとする。",
            "社会の仕組みや様々な職業の人に興味を持ち、敬意を払う。",
            "自分と他者の考えの違いを認め、相手を尊重しようとする。",
            "伝統的な行事に親しみ、地域社会との繋がりを感じる。",
            "家族の温かさを感じ、感謝の気持ちを言葉で表そうとする。"
        ],
        "環境": [
            "自然環境の保全に関心を持ち、自分にできることを考え行動する。",
            "動植物の命の尊さに気づき、愛情を持って育てようとする。",
            "数や量の概念を理解し、生活の中で測定や比較を楽しむ。",
            "文字や標識の機能に興味を持ち、情報として活用しようとする。",
            "科学的な事象（電気、磁石等）に触れ、その性質を探求する。",
            "地図や地球儀に興味を持ち、広い世界に関心を広げる。",
            "道具の安全な使い方を習得し、目的に合わせて正しく使う。",
            "カレンダーや時計を読み、計画的に活動を進めようとする。",
            "季節の変化を五感で捉え、美しさや不思議さを分かち合う。",
            "リサイクル活動に興味を持ち、物を大切に使い切ろうとする。"
        ],
        "言葉": [
            "相手の意図を汲み取り、状況に応じた言葉遣いを使い分ける。",
            "自分の意見を論理的に説明し、説得力を持って伝えようとする。",
            "読書を楽しみ、物語の世界に浸って多様な言葉を習得する。",
            "文字を読み書きすることに喜びを感じ、手紙交換等を楽しむ。",
            "言葉の響きや面白さを楽しみ、詩や物語を創作しようとする。",
            "話し合いの場において、司会や記録などの役割を経験する。",
            "分からない言葉を自分で調べたり、大人に聞いたりして解決する。",
            "言葉による自己表現を深め、自分の気持ちを正確に伝える。",
            "ユーモアのある表現を使い、会話を豊かに盛り上げる。",
            "他言語への関心を持ち、異なる文化の言葉に触れようとする。"
        ],
        "表現": [
            "多様な表現技法（スパッタリング、デカルコマニー等）を楽しむ。",
            "音楽の強弱やリズムを捉え、意図を持って楽器を演奏する。",
            "友達と協力して大型の制作物を作り、達成感を共有する。",
            "劇遊びにおいて、登場人物の心情を考えながら演じようとする。",
            "自分の経験や空想を、絵や文章を組み合わせて表現する。",
            "廃材を工夫して使い、動きのある動く玩具を作ろうとする。",
            "伝統的な芸術作品（絵画、陶芸等）に触れ、感性を磨く。",
            "自分の表現した作品の意図を、言葉で発表しようとする。",
            "友達の作品の良いところを具体的に指摘し、批評し合う。",
            "表現することを通して、自分自身の個性を発揮しようとする。"
        ]
    },
    "5歳児": {
        "健康": [
            "自分の健康を自分で守る意識を持ち、進んで健康管理を行う。",
            "集団生活の中で規律を保ち、健康的な生活リズムを自ら作る。",
            "難しい運動（縄跳び、跳び箱等）に粘り強く挑戦し、達成感を味わう。",
            "食事の礼儀作法を身につけ、感謝して食事を楽しもうとする。",
            "安全に対する判断力を養い、周囲の状況を見て適切に行動する。",
            "病気や怪我の予防について学び、自分や友達を労る。",
            "身の回りの整理整頓を徹底し、美しく整える習慣を身につける。",
            "心身の成長を自覚し、小学生になることへの期待を持つ。",
            "自分の体力を知り、活動の強度を調節しようとする。",
            "環境の変化に適応し、心身の安定を保とうと努める。"
        ],
        "人間関係": [
            "友達と力を合わせ、より大きな目的の達成を目指して行動する。",
            "民主的な話し合いを通して、集団のルールを自分たちで決める。",
            "互いの個性を認め合い、尊重し合う深い絆を築く。",
            "最高学年としての自覚を持ち、園全体のために進んで活動する。",
            "社会の決まりや公共心を理解し、責任ある行動を心がける。",
            "友達を信頼し、自分の弱みや悩みも打ち明けることができる。",
            "異なる意見を持つ相手とも対話し、合意形成を目指す。",
            "地域の人々やボランティアの方々と積極的に関わりを深める。",
            "命の尊厳や平和について考え、思いやりのある行動をとる。",
            "卒園に向けて感謝の気持ちを持ち、仲間との思い出を大切にする。"
        ],
        "環境": [
            "地球規模の環境問題に関心を持ち、環境保護意識を高める。",
            "生物のライフサイクルを理解し、命のつながりを感じ取る。",
            "論理的な思考を深め、予測を立てて実験や観察を楽しむ。",
            "文字や数を生活の便利な道具として、自在に使いこなす。",
            "世界の文化や歴史に興味を持ち、多様な価値観を学ぶ。",
            "IT機器やメディアの適切な活用法に触れ、情報を得る。",
            "時計を見て計画的に行動し、時間の管理を自分で行う。",
            "日本の伝統文化（茶道、書道等）に親しみ、その精神に触れる。",
            "数的な推論を楽しみ、図形の構成や分割を工夫する。",
            "身近な自然を科学的な視点で見つめ、発見を深める。"
        ],
        "言葉": [
            "豊かな語彙を使いこなし、ニュアンスの違う表現を楽しむ。",
            "長編の物語を読み、登場人物の心情や背景を深く理解する。",
            "話し合いにおいて、論点を整理し建設的な意見を述べる。",
            "自分の思いや考えを文章で綴り、自己表現を楽しむ。",
            "他者の話を共感を持って聞き、適切なアドバイスや助言をする。",
            "敬語などの丁寧な言葉遣いを、時と場合に応じて使い分ける。",
            "ニュースや時事問題に関心を持ち、言葉を通して世界を広げる。",
            "発表やスピーチを通して、自分の考えを堂々と人前で伝える。",
            "ユーモアや比喩を使いこなし、会話の質を高める。",
            "文字の読み書きをほぼ完成させ、就学に向けた準備を整える。"
        ],
        "表現": [
            "自分の内面や感情を、芸術的な活動を通して深く表現する。",
            "友達と合奏や合唱を創り上げ、調和する喜びを分かち合う。",
            "空間を意識した立体的な制作や、複雑な造形表現に挑む。",
            "劇や音楽発表において、演出や小道具を自分たちで工夫する。",
            "様々な芸術作品（名画、音楽、舞台）を鑑賞し、感性を養う。",
            "自分の作品をポートフォリオにまとめ、成長を振り返る。",
            "素材の特性を活かしきり、実用的な作品を完成させる。",
            "即興で踊ったり歌ったりして、自己を解放し表現を楽しむ。",
            "伝統工芸や郷土玩具の制作に触れ、手仕事の美しさを知る。",
            "自分自身の個性を確立し、オリジナリティ溢れる表現を追求する。"
        ]
    }
}
DEFAULT_TEXTS = ["（定型文を選択、または直接入力）", "自分で入力する"]

# --- 2. データベース操作関数 (保存・読込) ---

def load_data_from_sheet(user_id, doc_type):
    """スプレッドシートからデータを読み込み、セッションステートに反映する"""
    conn = st.connection("gsheets", type=GSheetsConnection)
    try:
        df = conn.read(ttl=0)
        # ユーザーIDと書類タイプで検索
        user_df = df[(df["user_id"] == user_id) & (df["doc_type"] == doc_type)]
        
        if not user_df.empty:
            # 最新のデータを取得
            latest_row = user_df.iloc[-1]
            json_str = latest_row["data_json"]
            data_dict = json.loads(json_str)
            
            # セッションステートに書き戻す
            for key, value in data_dict.items():
                # 日付型などの復元が必要な場合はここで処理可能だが、今回は文字列として戻す
                st.session_state[key] = value
            return True
        else:
            return False
    except Exception as e:
        st.error(f"読み込みエラー: {e}")
        return False

def save_data_to_sheet(user_id, doc_type):
    """現在のセッションステート（入力内容）をJSONにして保存する"""
    conn = st.connection("gsheets", type=GSheetsConnection)
    try:
        df = conn.read(ttl=0)
        
        # 保存対象のキーのみを抽出（ウィジェットのキーなど）
        save_dict = {}
        for key in st.session_state:
            # Streamlitの内部キーなどを除外して保存
            if isinstance(st.session_state[key], (str, int, float, bool, list)):
                save_dict[key] = st.session_state[key]
            # 日付型はJSONにできないので文字列変換
            elif isinstance(st.session_state[key], (datetime.date, datetime.datetime)):
                save_dict[key] = st.session_state[key].strftime("%Y-%m-%d")

        json_str = json.dumps(save_dict, ensure_ascii=False)
        now_str = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        
        # 新しい行を作成
        new_row = pd.DataFrame([{
            "user_id": user_id,
            "doc_type": doc_type,
            "updated_at": now_str,
            "data_json": json_str
        }])
        
        # 既存データがあれば、そのユーザー・タイプの古いデータを削除して上書きするロジックも可能だが、
        # ここではシンプルに「追記」して、読み込み時に「最新」を取る方式にする
        # (スプレッドシートが重くなる場合は、定期的に削除が必要)
        updated_df = pd.concat([df, new_row], ignore_index=True)
        conn.update(data=updated_df)
        return True
    except Exception as e:
        st.error(f"保存エラー: {e}")
        return False

# --- 3. Excel作成関数群 (前と同じなので省略せず記述) ---

def create_annual_excel(age, config, orientation):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"年間指導計画({age})"
    thin = Side(style='thin')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE if orientation == "横" else ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    
    # (レイアウト詳細は省略せず実装)
    ws.merge_cells("A1:C1")
    ws['A1'] = f"年間指導計画 ({age})"
    ws['A1'].font = Font(bold=True, size=16)
    
    row = 3
    fixed_items = [("年間目標", "年間目標"), ("健康・安全", "健康・安全")]
    for label, key in fixed_items:
        ws.merge_cells(f"A{row}:A{row+1}")
        ws.cell(row=row, column=1, value=label).fill = header_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row+1, column=1).border = border
        ws.merge_cells(f"B{row}:E{row+1}")
        c = ws.cell(row=row, column=2, value=config['values'].get(key, ""))
        c.alignment = top_left_align
        c.border = border
        row += 2

    # 4期メイン
    ws.cell(row=row, column=1, value="項目 / 期").fill = header_fill
    ws.cell(row=row, column=1).border = border
    for i, t_name in enumerate(TERMS):
        c = ws.cell(row=row, column=i+2, value=t_name)
        c.fill = header_fill
        c.border = border
    row += 1

    for item in config['mid_items']:
        ws.cell(row=row, column=1, value=item).fill = header_fill
        ws.cell(row=row, column=1).border = border
        for i, t_name in enumerate(TERMS):
            c = ws.cell(row=row, column=i+2, value=config['values'].get(f"{item}_{t_name}", ""))
            c.alignment = top_left_align
            c.border = border
        row += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def create_monthly_excel(age, target_month, config, num_weeks, orientation):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "指導計画表"
    thin = Side(style='thin')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    total_cols = 1 + num_weeks
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws['A1'] = f"【指導計画】 {target_month} ({age})"
    ws['A1'].font = Font(bold=True, size=14)
    
    row = 3
    # 簡易実装：主要データのみ出力
    ws.cell(row=row, column=1, value="項目").fill = header_fill
    for i in range(1, num_weeks+1):
        ws.cell(row=row, column=i+1, value=f"第{i}週").fill = header_fill
    row += 1
    
    mid_labels = [config[f'l_mid{r}'] for r in range(6, 16)]
    for label in mid_labels:
        ws.cell(row=row, column=1, value=label).fill = header_fill
        for w_idx in range(1, num_weeks + 1):
            key = f"{label}_週{w_idx}"
            ws.cell(row=row, column=w_idx+1, value=config['values'].get(key, "")).alignment = top_left_align
            ws.cell(row=row, column=w_idx+1).border = border
        row += 1
        
    output = BytesIO()
    wb.save(output)
    return output.getvalue()

def create_weekly_excel(age, config, orientation):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "週案"
    thin = Side(style='thin')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    header_fill = PatternFill(start_color="F2F2F2", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    top_left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)

    ws.merge_cells("A1:D1")
    ws['A1'] = f"【週案】 {config['week_range']} ({age})"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = "週のねらい"
    ws['B2'] = config['values'].get("weekly_aim", "")
    
    days = ["月", "火", "水", "木", "金", "土"]
    row_idx = 4
    for day in days:
        ws.cell(row=row_idx, column=1, value=day)
        ws.cell(row=row_idx, column=2, value=config['values'].get(f"activity_{day}", ""))
        row_idx += 1

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

# ▼▼▼ 修正後の万能AI関数 ▼▼▼
def ask_gemini_aim(age, keywords, doc_type="月間指導計画"):
    # SecretsからAPIキーを取得
    if "GEMINI_API_KEY" not in st.secrets:
        return "エラー: APIキーがSecretsに設定されていません。"
    
    api_key = st.secrets["GEMINI_API_KEY"]
    genai.configure(api_key=api_key)
    
    try:
        # モデル指定（2.5-flash）
        model = genai.GenerativeModel('models/gemini-2.5-flash')
        
        # 書類タイプによって命令文を変える
        if doc_type == "年間指導計画":
            target_desc = "1年間を通した長期的な「年間目標」"
        elif doc_type == "週案":
            target_desc = "1週間（月〜土）の短期的な「週のねらい」"
        else:
            target_desc = "1ヶ月間の「月間ねらい」"

        prompt = f"""
        あなたはベテラン保育士です。
        以下の条件で、{doc_type}における{target_desc}の文章を1つ作成してください。
        
        【条件】
        ・対象年齢: {age}
        ・キーワード: {keywords}
        ・文体: 保育の専門用語を用い、最後は「〜する。」などの言い切りで終える。
        ・文字数: 100文字〜150文字程度
        """
        
        response = model.generate_content(prompt)
        return response.text.strip()
            
    except Exception as e:
        return f"接続エラー: {str(e)}"
# ▲▲▲ 修正ここまで ▲▲▲


# --- 4. メイン画面構築 ---

# ロゴとタイトルの表示
col1, col2 = st.columns([1, 5])
with col1:
    try:
        st.image("logo.png", width=80) # ロゴ画像があれば表示
    except:
        st.write("📛") # 画像がない場合の代わり
with col2:
    st.title("保育指導計画システム")

# セッション初期化
if 'annual_data' not in st.session_state: st.session_state['annual_data'] = {}
if 'monthly_data' not in st.session_state: st.session_state['monthly_data'] = {}

# サイドバー設定
st.sidebar.header("⚙️ 設定")
age = st.sidebar.selectbox("対象年齢", ["0歳児", "1歳児", "2歳児", "3歳児", "4歳児", "5歳児"])
mode = st.sidebar.radio("作成する書類", ["年間指導計画", "月間指導計画", "週案"])
orient = st.sidebar.radio("用紙向き", ["横", "縦"])

# 掲示板へのリンク
st.sidebar.markdown("---")
st.sidebar.link_button("☕ 掲示板（休憩室）へ", "https://hoiku-bbs-ez5sr2ocp4ni2r4ypxuqx6.streamlit.app")
st.sidebar.markdown("---")

# 📥 データ保存・読込エリア（サイドバー下部）
st.sidebar.subheader("💾 データの保存・読込")
user_id = st.sidebar.text_input("先生のお名前 (ID)", placeholder="例: yamada")
st.sidebar.caption("名前を入力して保存すると、後で続きから始められます。")

c1, c2 = st.sidebar.columns(2)
if c1.button("データ保存"):
    if user_id:
        if save_data_to_sheet(user_id, mode):
            st.sidebar.success(f"{mode}を保存しました！")
    else:
        st.sidebar.error("名前を入力してください")

if c2.button("データ読込"):
    if user_id:
        if load_data_from_sheet(user_id, mode):
            st.sidebar.success("読み込みました！")
            st.rerun() # 画面を更新してデータを反映
        else:
            st.sidebar.warning("データが見つかりません")
    else:
        st.sidebar.error("名前を入力してください")


# ==========================================
# モードA：年間指導計画（修正版）
# ==========================================
if mode == "年間指導計画":
    st.header(f"📅 {age} 年間指導計画")

    # ▼ AIアシスタント（年間用）
    with st.expander("🤖 AIアシスタント（年間目標を作成）", expanded=True):
        c_ai1, c_ai2 = st.columns([3, 1])
        with c_ai1:
            ai_keywords = st.text_input("キーワード", placeholder="例：基本的生活習慣 信頼関係 自然との触れ合い")
        with c_ai2:
            if st.button("✨ 年間目標作成"):
                if ai_keywords:
                    with st.spinner("AIが思考中..."):
                        # doc_type="年間指導計画" を指定
                        gen_text = ask_gemini_aim(age, ai_keywords, doc_type="年間指導計画")
                        st.session_state["年間目標"] = gen_text # 保存用キーに直接入れる
                        st.success("作成しました！下の「年間目標」を確認してください。")
                else:
                    st.error("キーワードを入れてください")

    default_items = "園児の姿\nねらい\n養護（生命・情緒）\n教育（5領域）\n環境構成・援助\n保護者支援\n行事"
    mid_item_list = st.text_area("項目設定（改行区切り）", default_items).split('\n')

    user_values = {}
    t1, t2 = st.tabs(["📌 基本情報", "📝 各期の計画"])

    with t1:
        st.subheader("年間を通じた目標")
        # AIが作ったテキストが反映されるように session_state を活用
        user_values["年間目標"] = st.text_area("年間目標", key="年間目標", height=100)
        user_values["健康・安全"] = st.text_area("健康・安全・災害対策", key="健康・安全", height=100)

    with t2:
        # ▼ ここからプルダウン化の処理
        # 注意: TEIKEI_DATAにデータがないと選択肢が出ないので、
        # まだデータがない項目のために「自由入力」を必ず追加しています。
        cols = st.columns(4)
        
        # その年齢の定型文データを取得
        age_data = TEIKEI_DATA.get(age, {})
        
        for i, term in enumerate(TERMS):
            with cols[i]:
                st.markdown(f"**{term}**")
                for item in mid_item_list:
                    k = f"{item}_{term}"
                    
                    # 定型文があるかチェック
                    if item in age_data:
                        # 定型文がある場合 → プルダウン
                        options = age_data[item] + ["（自由入力）"]
                        val = st.selectbox(f"{item}", options, key=k)
                    else:
                        # 定型文がない場合 → テキストエリア（または「データなし」と表示してもOK）
                        # ここでは使い勝手のためテキストエリアを残しますが、完全プルダウン化したい場合は
                        # 空のリスト ["（選択肢なし）"] などを表示することになります
                        val = st.text_area(f"{item}", key=k, height=100)
                    
                    user_values[k] = val
                    
                    if term not in st.session_state['annual_data']: st.session_state['annual_data'][term] = {}
                    st.session_state['annual_data'][term][item] = val

   

    if st.button("🚀 Excel作成"):
        config = {'mid_items': mid_item_list, 'values': user_values}
        data = create_annual_excel(age, config, orient)
        st.download_button("📥 ダウンロード", data, f"年間計画_{age}.xlsx")
        # ▼▼▼ プレビュー機能 ▼▼▼
    st.markdown("---")
    st.subheader("👀 仕上がりプレビュー")
    
    with st.container(border=True):
        st.markdown("### 📅 年間指導計画表")
        
        # データを表形式（DataFrame）に変換して表示
        import pandas as pd
        
        # プレビュー用のデータを作る
        preview_data = {}
        for term in TERMS: # 1期, 2期...
            term_values = []
            for item in mid_item_list: # ねらい, 養護...
                # 入力された値を取り出す
                val = st.session_state.get('annual_data', {}).get(term, {}).get(item, "")
                term_values.append(val)
            preview_data[term] = term_values
            
        # 表を作成
        df_preview = pd.DataFrame(preview_data, index=mid_item_list)
        st.dataframe(df_preview, use_container_width=True)
    # ▲▲▲ プレビューここまで ▲▲▲

# ==========================================
# モードB：月間指導計画
# ==========================================
elif mode == "月間指導計画":
    st.header(f"📝 {age} 月間指導計画")
    # ▼▼▼ 追加コード：AIアシスタントエリア ▼▼▼
    with st.expander("🤖 AIアシスタント（キーワードから『ねらい』を作成）", expanded=True):
        c_ai1, c_ai2, c_ai3 = st.columns([2, 1, 1])
        with c_ai1:
            ai_keywords = st.text_input("キーワードを入力", placeholder="例：雪遊び 手袋 貸し借り 感染症予防")
        with c_ai2:
            target_week = st.selectbox("反映先", ["第1週", "第2週", "第3週", "第4週"])
        with c_ai3:
            st.write("") # レイアウト調整用
            if st.button("✨ AI作成"):
                if not ai_keywords:
                    st.error("キーワードを入れてください")
                else:
                    with st.spinner("AIが執筆中..."):
                        generated_text = ask_gemini_aim(age, ai_keywords)
                        
                        # 生成されたテキストを、対象の週の「ねらい」入力欄にセットする
                        # ※前回のコードで、ねらいのキーは "w{週番号}_6" となっていました
                        week_num = target_week.replace("第", "").replace("週", "") # "1", "2"...
                        target_key = f"w{week_num}_6"
                        
                        st.session_state[target_key] = generated_text
                        st.success(f"{target_week}の『ねらい』に入力しました！")
    # 日付などは保存対象外（毎回選択）とする運用がシンプル
    month_date = st.date_input("対象月", value=datetime.date.today())
    month_str = month_date.strftime("%Y年%m月")
    
    st.info("💡 年間計画のデータがあれば、ここから引用できます")
    if st.button("年間計画から引用"):
         # (連動ロジックは前のまま使用可能)
         pass

    num_weeks = 4
    l_mid = {r: st.text_input(f"項目{r}", val, key=f"lm_{r}") for r, val in zip(range(6, 16), ["ねらい", "養護", "教育", "環境", "支援", "行事", "連携", "食育", "健康", "その他"])}
    
    tabs = st.tabs([f"第{i}週" for i in range(1, 5)] + ["反省"])
    user_values = {}
    
    age_data = TEIKEI_DATA.get(age, {})
    
    for i in range(4):
        with tabs[i]:
            st.caption(f"第{i+1}週")
            for r_num, label in l_mid.items():
                # keyを一意にする: w(週)_(行番号)
                k = f"w{i+1}_{r_num}"
                
               # --- ここから修正 ---
                # 「ねらい」だけはAIや自由入力のために最初からテキストエリアにする
                if label == "ねらい":
                    val = st.text_area(label, key=k, height=100)
                # それ以外の項目で定型文がある場合はプルダウンにする
                elif label in age_data:
                    val = st.selectbox(label, age_data[label] + ["自由入力"], key=k)
                # 定型文がない項目はテキストエリア
                else:
                    val = st.text_area(label, key=k, height=60)
                # --- ここまで修正 ---
                user_values[f"{label}_週{i+1}"] = val
                
                if label == "ねらい":
                    st.session_state['monthly_data'][f"ねらい_週{i+1}"] = val

    with tabs[4]:
        user_values["reflection"] = st.text_area("振り返り", key="mon_ref", height=100)

    if st.button("🚀 Excel作成"):
        config = {**{f'l_mid{r}': val for r, val in l_mid.items()}, 'values': user_values}
        data = create_monthly_excel(age, month_str, config, num_weeks, orient)
        st.download_button("📥 ダウンロード", data, f"月案_{month_str}.xlsx")
        # ▼▼▼ プレビュー機能 ▼▼▼
    st.markdown("---")
    st.subheader("👀 仕上がりプレビュー")

    with st.container(border=True):
        st.markdown(f"### 🌙 {month_str} 指導計画")
        
        # 4週間分を並べて表示
        for i in range(4):
            week_num = i + 1
            with st.expander(f"第 {week_num} 週の内容を確認", expanded=True):
                # ユーザーが入力したデータを取得して表示
                w_aim = user_values.get(f"ねらい_週{week_num}", "（未入力）")
                st.markdown(f"**🎯 ねらい**: {w_aim}")
                
                # その他の項目をリストで表示
                for r_num, label in l_mid.items():
                    if label != "ねらい": # ねらい以外を表示
                        val = user_values.get(f"{label}_週{week_num}", "-")
                        if val:
                            st.text(f"【{label}】: {val}")
    # ▲▲▲ プレビューここまで ▲▲▲


# ==========================================
# モードC：週案（エラー対策強化版）
# ==========================================
elif mode == "週案":
    st.header(f"📅 {age} 週案")
    start_date = st.date_input("週の開始日")

    # セッションステートの初期化
    days = ["月", "火", "水", "木", "金", "土"]
    for d in days:
        for k in ["activity", "care", "tool"]:
            key_name = f"{k}_{d}"
            if key_name not in st.session_state:
                st.session_state[key_name] = ""

    # ▼ AI設定エリア
    with st.container(border=True):
        st.subheader("🤖 AI週案クリエイター")
        st.info("「今週のねらい」を入力してボタンを押すと、月〜土の計画を一括で提案します。")
        
        weekly_aim = st.text_area("今週のねらい（キーワードでもOK）", 
                                  key="weekly_aim_input", 
                                  height=80,
                                  placeholder="例：秋の自然に触れながら、戸外で体を動かして遊ぶ。")

        if st.button("✨ このねらいで1週間分を作成する"):
            if not weekly_aim:
                st.error("先に「ねらい」を入力してください。")
            else:
                with st.spinner("AIが6日分のカリキュラムを考案中..."):
                    try:
                        # プロンプト（AIへの命令文）
                        prompt = f"""
                        あなたはベテラン保育士です。以下の条件で週案を作成し、JSON形式のみを出力してください。
                        余計な挨拶やMarkdown記号（```json 等）は一切不要です。
                        
                        【条件】
                        ・対象年齢: {age}
                        ・今週のねらい: {weekly_aim}
                        ・月〜土の6日分
                        ・キーは必ず "月", "火", "水", "木", "金", "土" にする
                        
                        【出力データの例（この形式を守ること）】
                        {{
                            "月": {{"activity": "活動内容...", "care": "配慮...", "tool": "準備..."}},
                            "火": {{"activity": "...", "care": "...", "tool": "..."}}
                        }}
                        """
                        
                        # AI生成実行
                        model = genai.GenerativeModel('models/gemini-2.5-flash')
                        response = model.generate_content(prompt)
                        
                        # ▼▼▼ 修正ポイント：ここを強力にしました ▼▼▼
                        # AIの回答から { で始まり } で終わる部分だけを無理やり抜き出す
                        text_content = response.text
                        match = re.search(r'\{.*\}', text_content, re.DOTALL)
                        
                        if match:
                            json_str = match.group(0)
                            schedule_data = json.loads(json_str) # 変換
                            
                            # データの反映
                            for day_key, data_val in schedule_data.items():
                                if day_key in days:
                                    st.session_state[f"activity_{day_key}"] = data_val.get("activity", "")
                                    st.session_state[f"care_{day_key}"] = data_val.get("care", "")
                                    st.session_state[f"tool_{day_key}"] = data_val.get("tool", "")
                            
                            st.success("作成しました！下の欄を確認・修正してください。")
                            st.rerun()
                        else:
                            st.error("データの取得に失敗しました。もう一度ボタンを押してみてください。")
                            
                    except Exception as e:
                        # どんなエラーが出たか画面に表示する（デバッグ用）
                        st.error(f"エラーが発生しました: {e}")
                        st.text("▼AIからの返答（参考）")
                        st.code(response.text) # 原因究明のためにAIの返事を表示

    # ▼ 入力欄
    st.markdown("---")
    user_values = {}
    user_values["weekly_aim"] = weekly_aim 

    cols = st.columns(3)
    for i, day in enumerate(days):
        with cols[i%3]:
            st.subheader(f"{day}曜日")
            user_values[f"activity_{day}"] = st.text_area("活動", key=f"activity_{day}", height=100)
            user_values[f"care_{day}"] = st.text_area("配慮・援助", key=f"care_{day}", height=120)
            user_values[f"tool_{day}"] = st.text_area("準備", key=f"tool_{day}", height=60)

    # ▼ プレビューとExcel出力
    st.markdown("---")
    if st.button("🚀 Excel作成"):
        config = {'week_range': start_date.strftime('%Y/%m/%d〜'), 'values': user_values}
        data = create_weekly_excel(age, config, orient)
        
        # インデント修正済みのダウンロードボタン
        file_name = f"週案_{age}.xlsx" if 'age' in locals() else "週案_作成データ.xlsx"
        st.download_button("📥 ダウンロード", data, file_name)


        # ▼▼▼ プレビュー機能 ▼▼▼
    st.markdown("---")
    st.subheader("👀 仕上がりプレビュー")
    
    # 紙のような白い枠を作る
    with st.container(border=True):
        st.markdown(f"#### 📅 週のねらい")
        st.info(user_values.get("weekly_aim", "（未入力）"))
        
        st.markdown("#### 📅 日ごとの計画")
        # 3列で表示して見やすくする
        pv_cols = st.columns(3)
        days = ["月", "火", "水", "木", "金", "土"]
        
        for i, day in enumerate(days):
            with pv_cols[i % 3]:
                st.markdown(f"**【{day}曜日】**")
                # 内容があれば表示、なければ「-」
                act = user_values.get(f"activity_{day}", "-")
                care = user_values.get(f"care_{day}", "-")
                
                st.caption("▼活動")
                st.write(act if act else "（未入力）")
                st.caption("▼配慮")
                st.write(care if care else "（未入力）")
                st.divider() # 区切り線
    # ▲▲▲ プレビューここまで ▲▲▲
    














